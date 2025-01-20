import re
import time
import pandas as pd
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

# # ChromeDriver 경로 및 옵션 설정
# service = Service("chromedriver.exe")
# chrome_options = webdriver.ChromeOptions()
# chrome_options.add_argument("user-data-dir=C:/Temp/NewChromeProfile")  # 새 프로필 경로


options = webdriver.ChromeOptions()
options.add_argument('--no-sandbox')
options.add_argument('--disable-dev-shm-usage')
options.add_argument("user-data-dir=C:/Temp/NewChromeProfile")  # 새 프로필 경로


# WebDriver 초기화
driver = webdriver.Chrome(options=options)

# 서울 송파로 필터링(4~13단계)
def apply_filters1(driver):
    # 4~13단계 필터 적용
    driver.find_element(By.XPATH, '//*[@id="__next"]/div/div[2]/div[1]/div[1]/div/div[2]/div[2]/div/div/div/div/div[4]').click()
    driver.find_element(By.XPATH, '//*[@id="__next"]/div/div[2]/div[1]/div[1]/div/div[2]/div[2]/div/div[2]/div/div/div[18]').click()
    # driver.find_element(By.XPATH, '//*[@id="__next"]/div/div[2]/div[1]/div[2]/select[1]').click()
    # driver.find_element(By.XPATH, '//*[@id="__next"]/div/div[2]/div[1]/div[2]/select[1]/option[2]').click()
    driver.find_element(By.XPATH, '//*[@id="__next"]/div/div[2]/div[1]/div[2]/select[2]').click()
    driver.find_element(By.XPATH, '//*[@id="__next"]/div/div[2]/div[1]/div[2]/select[2]/option[2]').click()
    # driver.find_element(By.XPATH, '//*[@id="__next"]/div/div[2]/div[1]/div[2]/select[3]').click()
    # driver.find_element(By.XPATH, '//*[@id="__next"]/div/div[2]/div[1]/div[2]/select[3]/option[2]').click()
    driver.find_element(By.XPATH, '//*[@id="__next"]/div/div[2]/div[1]/div[2]/div/select').click()
    driver.find_element(By.XPATH, '//*[@id="__next"]/div/div[2]/div[1]/div[2]/div/select/option[2]').click()
    time.sleep(1)

# 서울 송파로 필터링(4~13단계)
def apply_filters2(driver):
    # 4~13단계 필터 적용
    driver.find_element(By.XPATH, '//*[@id="__next"]/div/div[2]/div[1]/div[1]/div/div[2]/div[2]/div/div[1]/div/div/div[5]').click()
    driver.find_element(By.XPATH, '//*[@id="__next"]/div/div[2]/div[1]/div[1]/div/div[2]/div[2]/div/div[2]/div/div/div[12]').click()
    # driver.find_element(By.XPATH, '//*[@id="__next"]/div/div[2]/div[1]/div[2]/select[1]').click()
    # driver.find_element(By.XPATH, '//*[@id="__next"]/div/div[2]/div[1]/div[2]/select[1]/option[2]').click()
    driver.find_element(By.XPATH, '//*[@id="__next"]/div/div[2]/div[1]/div[2]/select[2]').click()
    driver.find_element(By.XPATH, '//*[@id="__next"]/div/div[2]/div[1]/div[2]/select[2]/option[2]').click()
    # driver.find_element(By.XPATH, '//*[@id="__next"]/div/div[2]/div[1]/div[2]/select[3]').click()
    # driver.find_element(By.XPATH, '//*[@id="__next"]/div/div[2]/div[1]/div[2]/select[3]/option[2]').click()
    # driver.find_element(By.XPATH, '//*[@id="__next"]/div/div[2]/div[1]/div[2]/div/select').click()
    # driver.find_element(By.XPATH, '//*[@id="__next"]/div/div[2]/div[1]/div[2]/div/select/option[2]').click()
    time.sleep(1)

try:
    # 사이트 접속
    site_url = "https://www.reviewnote.co.kr/campaigns"
    driver.get(site_url)
    time.sleep(2)

    # 1. 로그인 버튼 확인
    try:
        login_button = driver.find_element(By.XPATH, '//*[@id="__next"]/div/div[1]/div[1]/div[2]/div[2]/a')
        if login_button.is_displayed():
            login_button.click()
            time.sleep(1)

            login_form = WebDriverWait(driver, 5).until(
                EC.presence_of_element_located((By.XPATH, '//*[@id="login"]'))
            )
            login_form.click()

            #
            time.sleep(20)
    except Exception:
        print("로그인 버튼 없음. 바로 크롤링 시작.")

    # 공통 작업 (4~13단계)
    # 캠페인 목록 및 필터링]
    apply_filters1(driver)
    time.sleep(1)

    # 크롤링 데이터 저장용 리스트
    data = []
    visited_links = set()  # 이미 방문한 링크를 저장
    scroll_attempts = 0  # 스크롤 시도 횟수
    last_height = driver.execute_script("return document.body.scrollHeight")  # 마지막 페이지 높이 추적

    # 크롤링 및 스크롤 루프
    while True:
        # 현재 페이지에서 아이템 요소를 다시 가져옴
        items = driver.find_elements(By.XPATH, '//*[@id="__next"]/div/div[2]/div[2]/div/div/div/a')

        # 새롭게 로드된 링크 추출
        links = [item.get_attribute("href") for item in items if item.get_attribute("href") not in visited_links]

        # 새 탭 열기
        driver.execute_script("window.open('');")
        driver.switch_to.window(driver.window_handles[-1])
        for link in links:
            try:
                visited_links.add(link)  # 방문한 링크 추가
                print(f"방문 중: {link}")

                driver.get(link)

                # 데이터 로드 대기
                WebDriverWait(driver, 5).until(
                    EC.presence_of_element_located((By.XPATH, '//*[@id="__next"]/div/div[2]/div/div[1]/div[1]/div[1]'))
                )

                # 데이터 크롤링
                store_name = driver.find_element(By.XPATH, '//*[@id="__next"]/div/div[2]/div/div[1]/div[1]/div[1]').text
                support = driver.execute_script(
                    "return document.evaluate('//*[@id=\"__next\"]/div/div[2]/div/div[1]/div[1]/div[3]/div[2]/div[5]/div[2]/span[1]', document, null, XPathResult.FIRST_ORDERED_NODE_TYPE, null).singleNodeValue.textContent;"
                )
                recruitment = driver.execute_script(
                    "return document.evaluate('//*[@id=\"__next\"]/div/div[2]/div/div[1]/div[1]/div[3]/div[2]/div[5]/div[2]/span[2]', document, null, XPathResult.FIRST_ORDERED_NODE_TYPE, null).singleNodeValue.textContent;"
                )
                service = driver.find_element(By.XPATH, '//*[@id="__next"]/div/div[2]/div/div[1]/div[1]/div[7]/div[1]/div[1]/div[2]/p').text
                address = driver.find_element(By.XPATH, '//*[@id="__next"]/div/div[2]/div/div[1]/div[1]/div[7]/div[1]/div[4]/div[2]').text
                visit_info = driver.find_element(By.XPATH, '//*[@id="__next"]/div/div[2]/div/div[1]/div[1]/div[7]/div[1]/div[6]/div[2]/p[1]').text
                period = driver.execute_script(
                    "return document.evaluate('//*[@id=\"__next\"]/div/div[2]/div/div[1]/div[1]/div[3]/div[2]/div[3]/div[2]', document, null, XPathResult.FIRST_ORDERED_NODE_TYPE, null).singleNodeValue.textContent;"
                )
                # 신청 여부 추가
                application_status = driver.execute_script(
                    "return document.evaluate('//*[@id=\"__next\"]/div/div[2]/div/div[1]/div[1]/div[3]/div[4]/button', document, null, XPathResult.FIRST_ORDERED_NODE_TYPE, null).singleNodeValue.textContent;"
                )
                # 데이터 저장
                data.append({
                    "가게 이름": store_name,
                    "지원": support,
                    "모집": recruitment,
                    "제공 서비스": service,
                    "주소": address,
                    "방문 및 예약 안내": visit_info,
                    "체험 기간": period,
                    "신청 여부": application_status,
                    "체험단 사이트 주소": link,  # 체험단 주소 추가
                })

            except Exception as e:
                print(f"크롤링 중 오류 발생 (링크: {link}): {e}")

        # 현재 탭 닫기 및 원래 탭으로 복귀
        driver.close()
        driver.switch_to.window(driver.window_handles[0])

        # 페이지를 아래로 스크롤
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(2)

        # 새로운 높이 확인
        new_height = driver.execute_script("return document.body.scrollHeight")
        if new_height == last_height:  # 더 이상 스크롤할 내용이 없음
            scroll_attempts += 1
        else:
            scroll_attempts = 0  # 새 데이터가 로드되었으므로 시도 횟수 초기화

        last_height = new_height

        if scroll_attempts >= 3:  # 3번 시도해도 새 데이터가 없으면 종료
            print("더 이상 새로운 항목이 없습니다.")
            break

    print("크롤링 완료!")
finally:
    driver.quit()

# 결과를 DataFrame으로 저장
df = pd.DataFrame(data)

# Excel 파일 저장 및 서식 설정
excel_file = "reviewnote_data.xlsx"
df.to_excel(excel_file, index=False, engine="openpyxl")

# Excel 서식 적용
wb = Workbook()
ws = wb.active

# 열 제목 작성 및 서식 적용
for col_num, column_title in enumerate(df.columns, 1):
    cell = ws.cell(row=1, column=col_num, value=column_title)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = Border(
        left=Side(style="thick"),
        right=Side(style="thick"),
        top=Side(style="thick"),
        bottom=Side(style="thick"),
    )

# 데이터 작성
for row_num, row_data in enumerate(df.values, 2):
    for col_num, cell_data in enumerate(row_data, 1):
        ws.cell(row=row_num, column=col_num, value=cell_data)

# 열 너비 자동 조정
for col in ws.columns:
    max_length = max(len(str(cell.value)) for cell in col if cell.value is not None)
    adjusted_width = max_length + 2
    ws.column_dimensions[col[0].column_letter].width = adjusted_width

# 저장
wb.save(excel_file)
print(f"크롤링 결과가 '{excel_file}'에 저장되었습니다.")
